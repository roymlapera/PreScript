import openpyxl
import numpy as np

def open_workbook(file_path, workbook_name):
    return openpyxl.load_workbook(file_path, read_only=True)[workbook_name]
    

def excel_column_list_generator():
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    column_list = []

    for letter1 in letters:
        column_list.append(letter1)

    for letter1 in letters:
        for letter2 in letters:
            column_list.append(letter1 + letter2)

    return column_list

def cell_data_importer(sheet,start_cell_idx, end_cell_idx, NUMERIC_VALUE=False):
    data = []

    excel_columns = excel_column_list_generator()
    
    for row in sheet.iter_rows(min_row= start_cell_idx[0], 
                               max_row= end_cell_idx[0], 
                               min_col= excel_columns.index(start_cell_idx[1])+1, 
                               max_col= excel_columns.index(end_cell_idx[1])+1, 
                               values_only=True):
        data.append(row)
    
    if NUMERIC_VALUE:
        return np.array(data, dtype=np.float32)  #casteando la lista a array de numpy de float32
    else:
        return np.array(data, dtype=str)  #casteando la lista a array de numpy de strings
    
def get_cell_content(file_path, cell_coordinate, sheet_name=None):
    try:
        # Open the Excel file
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        cell_content = []

        if sheet_name==None:
            # Select the desired sheet
            for sheet in workbook:
                # Get the content of the specified cell
                cell_content.append(sheet[cell_coordinate].value)
        else:
            sheet = workbook[sheet_name]
            # print(sheet[cell_coordinate])
            cell_content = sheet[cell_coordinate].value

        # Close the workbook
        workbook.close()
        return cell_content
    except Exception as e:
        print(f"Error: {e}")
        return None
    
def none_based_data_parser(data):

    if not isinstance(data, np.ndarray): data = np.array(data)

    # Create a boolean mask for rows where all values are 'None'
    none_rows_mask = np.all(data == 'None', axis=1)

    # Get indices of rows where all values are 'None'
    none_rows_indices = np.where(none_rows_mask)[0]

    # Split the array into chunks based on NaN rows
    chunks = np.split(data, none_rows_indices) if len(none_rows_indices) > 0 else [data]

    chunks = [np.array(chunk, dtype=str) for chunk in chunks if not np.all(chunk == 'None')]

    # print(f'Numero de chunks: {len(chunks)}')

    return chunks


